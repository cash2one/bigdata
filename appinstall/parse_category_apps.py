import time
import openpyxl
from tablib import Databook
from tablib import Dataset

import pymysql
import pymysql.cursors as cursors

connectConfig = {
    'host': '127.0.0.1',
    'port': 3306,
    'user': 'root',
    'password': 'root',
    'db': 'bigdata',
    'charset': 'utf8mb4',
    'cursorclass': cursors.DictCursor
}

phone_model_user_number = {
    "Y37": 100000,
    "X6D": 100000,
    "X6PlusD": 98243,
    "Xplay5A": 75305
}

phone_user_number_list = [100000,100000,98243,75305]

marker_apps_excel = "G:\\datasource\\refined_data\\appMarkers.xlsx"  # 定义输入appmarker的地址

input_data_path = "G:\\datasource\\appinstall_info"
date_list = "_".join(list(map(str, time.localtime()))[1:5])

excel_out_path = date_list + ".xlsx"  # 确定输出最终结果的的地址

user_packages = {}
package_users = {}

phone_model = ["Y37", "X6D", "X6PlusD", "Xplay5A"]

def load_app_category():
    """
    从数据库中加载category, 并整理成
    package:[app_name,main_category, secondary]
    :return:
    """
    dbConns = pymysql.connect(**connectConfig)
    with dbConns.cursor() as cursor:
        query = "select package, name ,maincategory, secondary from rawappcategory;"
        temp = cursor.execute(query)
        print(temp)
        raw_categorys = cursor.fetchmany(temp)
        categorys = {}
        for index in range(len(raw_categorys)):
            temp_line = raw_categorys[index]
            categorys[temp_line["package"]] = [temp_line["name"], temp_line["maincategory"], temp_line["secondary"]]
    return categorys


package_category = load_app_category()


def get_package_name(package):
    """
    输入包名，返回应用名称
    :param package:
    :return:
    """
    try:
        name = package_category[package][0]
        return name
    except KeyError as ex:
        print("key errors", "the input package is", package)
        try:
            return package_name[package]
        except KeyError:
            print("key errors  in the package_name map ", package)
            return "error"
        return "error"


def get_data_store_dir():
    """
    根据不同的操作系统返回不同的数据存放地址
    :return: user_packages和package_users的基础数据
    """
    import platform
    if "Windows" in platform.platform():
        return "G:\\datasource\\appinstall_info"
    else:
        return "Os X file path "


def load_data(phone_model):
    """
    根据机型名称，加载用户安装的包名，以及每个用户的
    :param phone_model:
    :return: None
    """
    temp_user_packages = {}
    with open(input_data_path + "\\"+ phone_model + "_user_packages.txt", "r") as f:
        for line in f:
            user_packages_line = line.split(",")
            temp_user_packages[user_packages_line[0]] = {
                "package_num": user_packages_line[1],
                "packages": user_packages_line[2:]
            }
    user_packages[phone_model] = temp_user_packages
    temp_package_users = {}
    with open(input_data_path+"\\"+ phone_model+"_package_users.txt", "r") as f:
        for line in f:
            package_users_line = line.split(",")
            temp_package_users[package_users_line[0]] = {
                "user_num": package_users_line[1],
                "user_list": package_users_line[2:]
            }

    package_users[phone_model] = temp_package_users


def get_target_package_name_map(excel_name):
    """
    输入excel，返回一个package 和 应用名称的dict
    :return:
    """
    package_name = {}
    try:
        work_book = openpyxl.load_workbook(excel_name)
        sheet_list = work_book.get_sheet_names()
        for sheet in sheet_list:
            try:
                sheet = work_book.get_sheet_by_name(sheet)
                package_cells = sheet.columns[0]
                name_cells = sheet.columns[1]
                for index in range(len(name_cells)):
                    package_name[package_cells[index].value] = name_cells[index].value
            except KeyError as ex:
                print("the input sheet name is:", sheet, "error happens", ex)
                return None
    except FileNotFoundError as e:
        print("the input file path is:", excel_name, "error happens:", e)
        return None

    return package_name


package_name = get_target_package_name_map(marker_apps_excel)


def get_wanted_packages(excel_name, sheet_name):
    """
    从目标Excel的sheet中提取目标package列表
    :param excel_name: excel文件名
    :param sheet_name: excel sheet名称
    :return: 所需要提取的目标package
    """
    package_list = []
    try:
        work_book = openpyxl.load_workbook(excel_name)
        sheet = work_book.get_sheet_by_name(sheet_name)
        package_cells = sheet.columns[0]
        for cell in package_cells:
            package_list.append(cell.value)
    except FileNotFoundError as e:
        print("the input file path is:", excel_name, "error happens:", e)
        return None
    except KeyError as ex:
        print("the input sheet name is:", sheet_name, "error happens", ex)
        return None
    return package_list


# 加载目标机型的所有package_users以及 user_packages数据
for model in phone_model:
    load_data(model)


def get_all_category(target_excel_path):
    """
    输入appmarker excel文件的目标地址，得到所有的categoryinfo
    :param target_excel_path: marker_app 存放的位置
    :return: 返回所有的目标app的分类，以及各个分类中包含的标记app， package的名称
    """
    import openpyxl
    category_info = {}
    try:
        work_book = openpyxl.load_workbook(target_excel_path)
    except FileNotFoundError:
        print("the wanted file ", target_excel_path, " is not found, please check")
        return None
    category_list = work_book.get_sheet_names()
    for category in category_list:
        package_list = []
        try:
            sheet = work_book.get_sheet_by_name(category)
            package_cells = sheet.columns[0]
            for cell in package_cells:
                package_list.append(cell.value)
        except KeyError as ex:
            print("error happens", ex)
            return None
        category_info[category] = package_list
    return category_info


def get_common_user(package_list, package_users):
    """
    设定packaglist，然后从package_users中得到获取这部分应用的共同用户数量
    :param package_users:
    :param package_list:
    :return:
    """
    if package_list is None or package_users is None:
        print("please check package_list and package_users")
    common_users = set()
    for package in package_list:
        try:
            temp_user_set = set(package_users[package]["user_list"])
            common_users &= temp_user_set
        except KeyError as key:
            print("could not find the package:" , package)
            return set()
    return common_users


def get_union_user(package_list, package_users):
    """
    输入目标的package，从package中获取这些package覆盖的用户数量
    :param package_lsit:
    :param package_users:
    :return:
    """
    if package_list is None or package_users is None:
        print("please check package_list and package_users")
    union_users = set()
    for package in package_list:
        try:
            temp_user_set = set(package_users[package]["user_list"])
            union_users |= temp_user_set
        except KeyError as key:
            print("could not find the package:" , package)
            continue
    return union_users


def get_whole_results(output_dir_path="./"):
    """
    根据输入的机型数据和 输出类型 输出计算结果，默认的输出类型为excel文件，但是后期可以更换为json等，来满足网页显示
    不需要输出phone的名称，这个函数返回最终的计算结果，并以excel的形式展现
    :param output_dir_path: 存放文件的目录
    :param phone_model: 机型名称
    :param output_type: 输出结果的类型
    :return: None
    """
    work_book = Databook()
    category_info = get_all_category(marker_apps_excel)
    common_headers = []
    for model in phone_model:
        common_headers.append(model)

    summary = {}
    for category in category_info:
        category_headers = list(common_headers)
        category_headers.insert(0, category)
        data_list = []
        package_list = category_info[category]
        for package in package_list:
            package_data = [get_package_name(package)]
            for model in phone_model:
                try:
                    user_number = package_users[model][package]["user_num"]
                except KeyError as ke:
                    user_number = 0
                package_data.append(str(user_number))
            data_list.append(tuple(package_data))
        unique_user = ["unique user"]
        print(data_list)
        for model in phone_model:
            unique_user_number = len(get_union_user(package_list,  package_users[model]))
            unique_user.append(unique_user_number)
        summary[category]  = unique_user[1:]
        data_list.append(tuple(unique_user))
        work_sheet = Dataset(*data_list, title=category, headers=category_headers)
        work_book.add_sheet(work_sheet)

    summary_data = []
    for category in summary:
        user_numbers = summary[category]
        rate_line = [category]
        for index in range(len(user_numbers)):
            rate_line.append(int(user_numbers[index])/phone_user_number_list[index])
        # for model in phone_model:
        #     rate_line.append(int(user_number)/phone_model_user_number[model])
        summary_data.append(tuple(rate_line))

    summary_headers = ["分类"]
    for model in phone_model:
        summary_headers.append(model)

    import  time
    summary_sheet  = Dataset(*summary_data,headers= summary_headers, title="summary" )
    with open(output_dir_path + "summary.xlsx","wb") as f:
        f.write(summary_sheet.xlsx)

    with open(output_dir_path + "details.xlsx", "wb") as  f:
        f.write(work_book.xlsx)


get_whole_results()
