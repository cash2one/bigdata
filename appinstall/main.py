import pymysql
import pymysql.cursors as cursors
import csv
import sys
import collections
import openpyxl
import time
from tablib import Databook
from tablib import Dataset

connectConfig = {
    'host': '127.0.0.1',
    'port': 3306,
    'user': 'root',
    'password': 'root',
    'db': 'bigdata2',
    'charset': 'utf8mb4',
    'cursorclass': cursors.DictCursor
    }
# 全局变量
isDebug = True             # 调试模式下只提取表中的1000个数据；
dataConfig = [{'model': "x6d", 'table': "x6app"}, {'model': 'x6plus', 'table': "x6plusapp"},
              {'model': "xplay5", "table": "xplay5app"}, {'model': "y37", 'table': "y37app"}]
excel_in_path = "G:\\datasource\\refined_data\\appMarkers.xlsx"         # 定义输入appmarker的地址
date_list = "_".join(list(map(str, time.localtime()))[1:5])
excel_out_path = "G:\\bigdata\\result\\app\\appinstall_" + date_list + ".xlsx"            # 确定输出最终结果的的地址
table_content = {}                              # 用来保存各个数据库上的应用列表
user_app = []           # 对于不同的机型 存储每个用户对应的app列表
app_user = []           # 对于不同的机型 存储每个app安装的用户列表


def date_process_init():
    """
    从数据库中读取数据原始数据
    :return:
    """
    for data in dataConfig:
        if isDebug:
            table_content[data["model"]] = get_table_content(data["table"], 1000)
        else:
            table_content[data["model"]] = get_table_content(data["table"])
    return True


def get_user_app(content = table_content):
    """
    解析得到 每个用户安装的应用列表
    :param content: sql 查询返回的类别
    :return:
    """
    result = []
    return result


def get_app_user(content = table_content):
    """
    解析每个app被哪些用户安装
    :param content: sql 查询返回的列表
    :return:
    """
    result = []
    return result


def get_table_content(table_name, row_num="all"):
    """
    输入目标数据表名称，返回数据表中对应的内容
    :param table_name: 要查询的数据表
    :param row_num: 要查询行数量，不传入数据时返回所有的行。
    :return: 所查询到的内容
    """
    db_connection = pymysql.connect(**connectConfig)
    if row_num is not "all":
        sql = "select * from " + table_name + " limit " + str(row_num)
    else:
        sql = "select * from " + table_name
    try:
        with db_connection.cursor() as cursor:
            temp = cursor.execute(sql)
            result = cursor.fetchmany(temp)
            return result
    except Exception as ex:
        print("exception happens while trying to access table content", ex)
        return None
    finally:
        db_connection.close()


def init_distinct_user_dict():
    distinct_user = {}
    for data in dataConfig:
       distinct_user[data["model"]] = 0;
    return distinct_user


def get_distinct_user_num(model = "all"):
    if model == "all":
        return distinct_user
    else:
        try:
            user_num = distinct_user["model"]
        except KeyError:
            return "输入的机型名称为"+ model + ", 查找不到该机型的数据，请确认机型名称是否有错误"
    return user_num
distinct_user = init_distinct_user_dict()
# date_process_init()


def get_imei_from_package(table_name , package_name):
    """
    输入对应的表名，和文件名，返回imei的列表
    :param table_name: 对应的table名称
    :param package_name: 对应的包名
    :return: user数量
    """
    imei_list = []
    db_connection = pymysql.connect(**connectConfig)
    try:
        SQL = "select imei from "+ table_name + " where package like \"%" + package_name + "%\""
        with db_connection.cursor() as cursor:
            temp = cursor.execute(SQL)
            result = cursor.fetchmany(temp)
        for element in result:
            imei_list.append(element["imei"])
    except Exception as ex:
        print(ex)
    finally:
        db_connection.close()
    return imei_list


def get_wanted_packages(excel_name, sheet_name):
    """
    从目标Excel的sheet中提取目标package列表
    :param excel_name: excel文件名
    :param sheet_name: excel sheet名称
    :return: 所需要提取的目标package
    """
    package_list = []
    try:
        work_book = openpyxl.load_workbook(excel_name)
        sheet = work_book.get_sheet_by_name(sheet_name)
        package_cells = sheet.columns[0]
        for cell in package_cells:
            package_list.append(cell.value)
    except FileNotFoundError as e:
        print("the input file path is:", excel_name, "error happens:", e)
        return None
    except KeyError as ex:
        print("the input sheet name is:", sheet_name, "error happens", ex)
        return None
    return package_list

# 根据输入的
wanted_package_category_work_book = openpyxl.load_workbook(excel_in_path)
wanted_package_category_list = wanted_package_category_work_book.get_sheet_names()

model_list = ["x6app", "x6plusapp", "y37app", "xplay5app"]      # hardcode 需要进一步更改
general_result = Databook()                                             # 记录所有的机型和app类别安装信息
for model in model_list:
    model_general_result = Dataset()  # 结果的第一个汇总值的内容

    for package_category in wanted_package_category_list:
        general_category_
        wanted_packages = get_wanted_packages(excel_in_path, package_category)
        imei_set = set()
        package_count = 0
        model_specific_category_result = {}  # 记录特定类别中单独应用的使用人数
        for package in wanted_packages:
            package_count += 1
            print(model, "package category is ", package_category, "percent:", package_count / len(wanted_packages))
            imei_list = get_imei_from_package(model, package)
            imei_set = set(imei_list) | imei_set                # 记录特定类别的整体使用人数
            model_specific_category_result[package] = len(imei_list)
        model_specific_category_result["general"] = len(imei_set)
        print("model:", model, package_category, "应用数量", len(imei_set))
        model_general_result[package_category] = model_specific_category_result
    general_result[model] = model_general_result


##　临时代码　

# wanted_package_category_work_book = openpyxl.load_workbook(excel_in_path)
# wanted_package_category_list = wanted_package_category_work_book.get_sheet_names()
#
# model_list = ["x6app", "x6plusapp", "y37app", "xplay5app"]      # hardcode 需要进一步更改
# general_result = {}                                             # 记录所有的机型和app类别安装信息
# for model in model_list:
#     model_general_result = {}  # 结果的第一个汇总值的内容
#     result_workbook = openpyxl.Workbook()
#     package_category = "直播"
#     wanted_packages = get_wanted_packages(excel_in_path, package_category)
#     print(wanted_packages)
#     imei_set = set()
#     package_count = 0
#     model_specific_category_result = {}  # 记录特定类别中单独应用的使用人数
#     for package in wanted_packages:
#         package_count += 1
#         print(model, "package category is ", package_category, "percent:", package_count / len(wanted_packages))
#         imei_list = get_imei_from_package(model, package)
#         imei_set = set(imei_list) | imei_set                # 记录特定类别的整体使用人数
#         model_specific_category_result[package] = len(imei_list)
#     model_specific_category_result["general"] = len(imei_set)
#     print("model:", model, package_category, "应用数量", len(imei_set))
#     model_general_result[package_category] = model_specific_category_result
#     general_result[model] = model_general_result


def write_to_xlsx(content, dir_path, file_path="example"):
    """
    将content中的内容保存到目标文件中，content的类型有两种：
    dataset: 相当于单个sheet文件，采用
    :param content: 输入的content模式要满足一定的格式要求，是一个sheet_list, 该list中的每个元素对应到content中的
    :param file_path: 文件夹内容
    :return:
    """
    suffix = ".xlsx"
    hour_minute = time.localtime().tm_hour + "_"  + time.localtime().tm_min
    # 检查格式是否合适
    if type(content) is Dataset:
        file_name = content.title + hour_minute
    elif type(content) is Databook:
        file_name = file_path + hour_minute
    else:
        print("the input content is not Dataset or Databook, please check the input data format")
        return
    with open(dir_path + file_name + suffix, "wb") as f:
        f.write(content.xlsx)


def parse_command(input_command ="help"):
    """
    解析输入的结果
    :param input_command:
    :return: 如果命令解析成功 则返回success，否则返回命令解析失败和对应的error_code
    """
    help_str = """
    1. 独立用户数量 需再次输入机型名称；
    2. 特定应用在 各个机型上的安装数量和比例；
    3. 一串包名 在各个机型上的独立用户数量和比例
    4. 查看内置特定类别应用
    exit: 退出
    """
    if input_command == "help":
        return help_str
    elif input_command == '1':
        return ""

while True:
    command = input('请输入要执行的命令，输入help查看帮助，退出请输入exit;')
    command_striped = command.rstrip().strip("\"")
    if command_striped == "exit":
        print("程序退出")
        break
    parse_result = parse_command(command_striped)
    print(parse_result)


